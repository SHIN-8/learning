import openpyxl

file_name = input("xlsxファイル名を入力してください。")

#xlsxファイルを開く
wb = openpyxl.load_workbook(filename=file_name)
ws = wb.active

count  = 0        # 99の件数をカウントする変数
total_kingaku = 0  # 平均事故金額を保存する変数
max_kingaku = 0      # 最大事故金額を保存する変数
min_kingaku = 0      # 最小事故金額を保存する変数
row = 2              # 事故金額を取得する行


while row < 500:
    #トラッカーが99か判定する
    tracker = ws.cell(row=row, column=2).value
    if not tracker == "【99】":
        row = row + 1
        continue

    ## 99の件数を増やす
    count = count + 1

    #5列目に事故金額のセルがあるので、行を更新しながら事故金額を取得する
    value = ws.cell(row=row, column=5).value

    #valueは文字列なので整数に変換
    kingaku = int(value)

    #合計事故金額を計算する
    total_kingaku = total_kingaku + kingaku

    #最大事故金額より大きいか判定
    if max_kingaku < kingaku:
        max_kingaku = kingaku
    #最小事故金額より大きいか判定
    if min_kingaku > kingaku:
        min_kingaku = kingaku
    row = row + 1
    

#コンソールに表示する
print("99の件数は", count, "件です")
print("平均事故金額は", total_kingaku / count)
print("最大事故金額は", max_kingaku)
print("最小事故金額は", min_kingaku)